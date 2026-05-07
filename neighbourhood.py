# Python module for the TRNSYS Type calling Python using CFFI
# Data exchange with TRNSYS uses a dictionary, called TRNData in this file (it is the argument of all functions).
# Data for this module will be in a nested dictionary under the module name,
# i.e. if this file is calle "MyScript.py", the inputs will be in TRNData["MyScript"]["inputs"]
# for convenience the module name is saved in thisModule
#
# MKu, 2022-02-15

import math
import os
import pygfunction as gt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import minimize_scalar

from juliacall import Main as jl
from juliacall import Pkg as jlPkg

# Initialize Julia (once at DLL startup)
jl.seval("""
using GSHPsDesigner
""")
FLS_IV = jl.GSHPsDesigner.FLS_IV

thisModule = os.path.splitext(os.path.basename(__file__))[0]

# Initialization: function called at TRNSYS initialization
# ---------------------------------------------------------------------------------------------------------------------
def Initialization(TRNData):
    global borefield
    global LoadAgg
    global H
    global Q_tot
    global objective_function
    global m_flow_network
    global T_g
    global Rb
    global cp_f
    global Tf_in
    global Tf_out
    global dt
    global dT_n
    global index_year
    global dT_neighbourhood
    global dT_n
    global dT_neighbourhood_dt
    global nSteps
    
    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx", data_only=True)
    sheet = wb['Borehole']

    H = float(sheet[f"A2"].value)
    D = float(sheet[f"B2"].value)  
    rb = float(sheet[f"C2"].value)  
    x = float(sheet[f"D2"].value)
    y = float(sheet[f"E2"].value)   
    activation_year = float(sheet[f"F2"].value)  

    new_borehole_characteristics = [
        activation_year,            
        x,
        y,
        D,
        H,
        0],
        
    new_borehole_characteristics = np.array(new_borehole_characteristics)
  
    # Create borehole object
    borefield = [gt.boreholes.Borehole(H, D, rb, x, y)]

    # Load the ground properties
    sheet = wb['Ground']
    k = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    T_g = float(sheet['E2'].value)
    Rb = float(sheet['F2'].value)

    # Fluid properties
    sheet = wb['Fluid']
    cp_f = float(sheet['A2'].value)

    nSteps = TRNData[thisModule]["total number of time steps"]

    # Simulation parameters (must be consistent with TRNSYS!)
    dt = TRNData[thisModule]["simulation time step"] * 3600.
    tmax = nSteps * dt
    Nt = int(np.ceil(tmax/dt))
    time = dt * np.arange(1,Nt+1)

    LoadAgg = gt.load_aggregation.ClaessonJaved(dt,tmax)
    time_req = LoadAgg.get_times_for_simulation()

    gFunc = gt.gfunction.gFunction(borefield, k/(rho*cp), time=time_req)
    LoadAgg.initialize(gFunc.gFunc/(2*np.pi*k))

    Q_tot = np.zeros(nSteps) + H*10
    Tf_in = np.zeros(nSteps)
    Tf_out = np.zeros(nSteps)

    # Load the neighbourhood properties

    sheet = wb['NeighbouringBoreholes']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1

    # Number of boreholes
    Nb = filled_rows - 1

    Hn_list = [None] * (Nb)
    Dn_list = [None] * (Nb)
    extraction_list = [None] * (Nb)
    xn_list = [None] * (Nb)
    yn_list = [None] * (Nb)
    activation_year_nlist = [None] * (Nb)
    field_characteristics = []

    for row in range(2,filled_rows + 1):
        Hn_list[row-2] = float(sheet[f"A{row}"].value)
        Dn_list[row-2] = float(sheet[f"B{row}"].value)  
        xn_list[row-2] = float(sheet[f"D{row}"].value)
        yn_list[row-2] = float(sheet[f"E{row}"].value)   
        extraction_list[row-2] = float(sheet[f"F{row}"].value)  
        activation_year_nlist[row-2] = float(sheet[f"G{row}"].value)  
        field_characteristics.append([
        activation_year_nlist[row-2],
        xn_list[row-2],
        yn_list[row-2],
        Dn_list[row-2],
        Hn_list[row-2],
        extraction_list[row-2],
        ])

    field_characteristics = np.array(field_characteristics)

    if activation_year < max(activation_year_nlist):
        raise ValueError(f"The borehole activation year ({activation_year}) cannot be earlier than neighbouring boreholes (earliest: {max(activation_year_nlist)}). Please adjust the input data.")

    time_neighbourhood = math.floor(nSteps/8760); 

    results  = FLS_IV(jl.Array(field_characteristics), jl.Array(new_borehole_characteristics), time_neighbourhood, rb=7.5e-2, T_ground = T_g, k=k, ro=rho, cp=cp)

    if results[1].shape[1] > time_neighbourhood:
        dT_neighbourhood = T_g - np.array(results[1][-1,-time_neighbourhood-1:])
    else:
        dT_neighbourhood = T_g - np.array(results[1][-1,-time_neighbourhood:])
        dT_neighbourhood = np.concatenate(([0], dT_neighbourhood))



    index_year = 0
    dT_n = dT_neighbourhood[index_year]

    xp = np.linspace(0, 1, len(dT_neighbourhood))      # original positions
    x_new = np.linspace(0, 1, nSteps)       # new positions

    dT_neighbourhood_dt = np.interp(x_new, xp, dT_neighbourhood)

    def objective_function(x, T_in, dT_n, m_flow_network, cp_f, T_g, LoadAgg, H, Rb):
        # # x is the total load [W]    
        LoadAgg.set_current_load(x/H)
        deltaT_b = LoadAgg.temporal_superposition()
        T_b = T_g - deltaT_b - dT_n

        Tf = T_b - x/H * Rb
        T_f_in_single = Tf - ( x/2/m_flow_network/cp_f)
        
        return abs(T_f_in_single - T_in)

    return


# StartTime: function called at TRNSYS starting time (not an actual time step, initial values should be reported)
# ----------------------------------------------------------------------------------------------------------------------
def StartTime(TRNData):
    # with open("Result.txt","w") as file:
    #     pass
    # np.savetxt("dT_neighbourhood_TR.txt", dT_neighbourhood)
    # np.savetxt("nSteps.txt", np.array(nSteps))
    return

# Iteration: function called at each TRNSYS iteration within a time step
# ----------------------------------------------------------------------------------------------------------------------
def Iteration(TRNData):

    Tin = TRNData[thisModule]["inputs"][0]
    m_flow_network = TRNData[thisModule]["inputs"][1]
    stepNo = TRNData[thisModule]["current time step number"]

    LoadAgg.next_time_step(stepNo * dt)
    # if stepNo ==1:
    #     dT_n = dT_neighbourhood[0]
    # elif stepNo % 8760 ==0:
    #     index_year += 1
    #     dT_n = dT_neighbourhood[index_year]

    dT_n = dT_neighbourhood_dt[stepNo-1]
    
    solution = minimize_scalar(objective_function, args = (Tin, dT_n, m_flow_network, cp_f, T_g, LoadAgg, H,Rb),  method='brent')
    Q_tot[stepNo-1] = solution.x

    LoadAgg.set_current_load(Q_tot[stepNo-1] /H)
    deltaT_b = LoadAgg.temporal_superposition()
    T_b = T_g - deltaT_b - dT_n

    Tf = T_b - Q_tot[stepNo-1]/H*Rb
    Tf_in[stepNo -1] = Tf - ( Q_tot[stepNo-1]/2/m_flow_network/cp_f)

    Tf_out[stepNo -1] = Tf + ( Q_tot[stepNo-1]/2/m_flow_network/cp_f)


    # Set outputs in TRNData
    TRNData[thisModule]["outputs"][0] = Tf_out[stepNo -1]
    TRNData[thisModule]["outputs"][1] = Q_tot[stepNo -1]

    # with open("Result.txt","a") as file:
    #     file.write(str(dT_n)+"\n")
    #     pass

    return

# EndOfTimeStep: function called at the end of each time step, after iteration and before moving on to next time step
# ----------------------------------------------------------------------------------------------------------------------
def EndOfTimeStep(TRNData):

    # This model has nothing to do during the end-of-step call
    
    return


# LastCallOfSimulation: function called at the end of the simulation (once) - outputs are meaningless at this call
# ----------------------------------------------------------------------------------------------------------------------
def LastCallOfSimulation(TRNData):

    # NOTE: TRNSYS performs this call AFTER the executable (the online plotter if there is one) is closed. 
    # Python errors in this function will be difficult (or impossible) to diagnose as they will produce no message.
    # A recommended alternative for "end of simulation" actions it to implement them in the EndOfTimeStep() part, 
    # within a condition that the last time step has been reached.
    #
    # Example (to be placed in EndOfTimeStep()):
    #
    # stepNo = TRNData[thisModule]["current time step number"]
    # nSteps = TRNData[thisModule]["total number of time steps"]
    # if stepNo == nSteps-1:     # Remember: TRNSYS steps go from 0 to (number of steps - 1)
    #     do stuff that needs to be done only at the end of simulation

    return