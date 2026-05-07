import subprocess           # to run the TRNSYS simulation
import shutil               # to duplicate the output txt file
import time                 # to measure the computation time
import pytest               # for testing
import pandas as pd
import numpy as np
from numpy.testing import assert_array_equal
from openpyxl import load_workbook
import pygfunction as gt
from scipy.optimize import minimize_scalar
import os
import math

from juliacall import Main as jl
from juliacall import Pkg as jlPkg

jl.seval("""
using GSHPsDesigner
""")
FLS_IV = jl.GSHPsDesigner.FLS_IV

### The following parameters need to be tuned with the TRNSYS simulation file
nSteps = 17520
dt = 3600.
m_flow_network = 1.2
Tin = -3
############################################################################

def trnsys_results():
    deck_file_name = 'neighbourhood.dck'
    
    subprocess.run([r"C:\Trnsys18\Exe\TRNExe64.exe",r"C:\TRNSYS18\TRNLib\CallingPython-Cffi\Examples\trnsys-neighbourhood\neighbourhood.dck","/h"])

def objective_function(x, T_in, dT_n, m_flow_network, cp_f, T_g, LoadAgg, H, Rb):
    # # x is the total load [W]    
    LoadAgg.set_current_load(x/H)
    deltaT_b = LoadAgg.temporal_superposition()
    T_b = T_g - deltaT_b - dT_n

    Tf = T_b - x/H * Rb
    T_f_in_single = Tf - ( x/2/m_flow_network/cp_f)
    
    return abs(T_f_in_single - T_in)

def python_results():
    global nSteps, dt, m_flow_network, Tin

    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx", data_only=True)
    sheet = wb['Borehole']

    H = float(sheet[f"A2"].value)
    D = float(sheet[f"B2"].value)  
    rb = float(sheet[f"C2"].value)  
    x = float(sheet[f"D2"].value)
    y = float(sheet[f"E2"].value)   
    activation_year = float(sheet[f"F2"].value)  

    new_borehole_characteristics = [
        activation_year,            
        x,
        y,
        D,
        H,
        0],
        
    new_borehole_characteristics = np.array(new_borehole_characteristics)
  
    # Create borehole object
    borefield = [gt.boreholes.Borehole(H, D, rb, x, y)]

    # Load the ground properties
    sheet = wb['Ground']
    k = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    T_g = float(sheet['E2'].value)
    Rb = float(sheet['F2'].value)

    # Fluid properties
    sheet = wb['Fluid']
    cp_f = float(sheet['A2'].value)

    tmax = nSteps * dt

    Nt = int(np.ceil(tmax/dt))
    time = dt * np.arange(1,Nt+1)

    LoadAgg = gt.load_aggregation.ClaessonJaved(dt,tmax)
    time_req = LoadAgg.get_times_for_simulation()

    gFunc = gt.gfunction.gFunction(borefield, k/rho/cp, time=time_req)
    LoadAgg.initialize(gFunc.gFunc/(2*np.pi*k))

    dT = np.zeros(Nt)
    Tf_in = np.zeros(Nt)
    Tf_out = np.zeros(Nt)
    Q_tot = np.zeros(Nt) + H*10

    # Load the neighbourhood properties

    sheet = wb['NeighbouringBoreholes']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1

    # Number of boreholes
    Nb = filled_rows - 1

    Hn_list = [None] * (Nb)
    Dn_list = [None] * (Nb)
    extraction_list = [None] * (Nb)
    xn_list = [None] * (Nb)
    yn_list = [None] * (Nb)
    activation_year_nlist = [None] * (Nb)
    field_characteristics = []

    for row in range(2,filled_rows + 1):
        Hn_list[row-2] = float(sheet[f"A{row}"].value)
        Dn_list[row-2] = float(sheet[f"B{row}"].value)  
        xn_list[row-2] = float(sheet[f"D{row}"].value)
        yn_list[row-2] = float(sheet[f"E{row}"].value)   
        extraction_list[row-2] = float(sheet[f"F{row}"].value)  
        activation_year_nlist[row-2] = float(sheet[f"G{row}"].value)  
        field_characteristics.append([
        activation_year_nlist[row-2],
        xn_list[row-2],
        yn_list[row-2],
        Dn_list[row-2],
        Hn_list[row-2],
        extraction_list[row-2],
        ])

    field_characteristics = np.array(field_characteristics)

    time_neighbourhood = math.ceil(Nt/8760); 

    results  = FLS_IV(jl.Array(field_characteristics), jl.Array(new_borehole_characteristics), time_neighbourhood, rb=7.5e-2, T_ground = T_g, k=k, ro=rho, cp=cp)

    if results[1].shape[1] > time_neighbourhood:
        dT_neighbourhood = T_g - np.array(results[1][-1,-time_neighbourhood-1:])
    else:
        dT_neighbourhood = T_g - np.array(results[1][-1,-time_neighbourhood:])
        dT_neighbourhood = np.concatenate(([0], dT_neighbourhood))

    # np.savetxt("dT_neighbourhood_PY.txt", dT_neighbourhood)
    xp = np.linspace(0, 1, len(dT_neighbourhood))      # original positions
    x_new = np.linspace(0, 1, Nt)       # new positions

    dT_neighbourhood_dt = np.interp(x_new, xp, dT_neighbourhood) 

    # with open("dT_neighbourhood_dt.txt","w") as file:

    #     pass

    for i in range(0,Nt):

        LoadAgg.next_time_step(i * dt)
        dT_n = dT_neighbourhood_dt[i]  

        solution = minimize_scalar(objective_function, args = (Tin, dT_n, m_flow_network, cp_f, T_g, LoadAgg, H,Rb), method='brent')

        Q_tot[i] = solution.x

        LoadAgg.set_current_load(Q_tot[i] /H)
        deltaT_b = LoadAgg.temporal_superposition()

        T_b = T_g - deltaT_b - dT_n

        Tf = T_b - Q_tot[i]/H*Rb
        Tf_in[i] = Tf.item() - ( Q_tot[i].item()/2/m_flow_network/cp_f)

        Tf_out[i] = Tf.item() + ( Q_tot[i].item()/2/m_flow_network/cp_f)
 
        # with open("dT_neighbourhood_dt.txt","a") as file:
        #     file.write(str(dT_n)+"\n")
        #     pass


    return Tf_out

def test_neighbourhood():
    
    # Create results with trnsys
    trnsys_results()
    results_trnsys = pd.read_csv('Tout_neighbourhood.txt', sep=r'\s+', skiprows=1, names=["TIME", "Tout"])
    T_trnsys = np.array(results_trnsys['Tout'])

    # Create results with pygfunction
    T_python = python_results()

    # Compare
    np.testing.assert_allclose(T_python[0:8760*2], T_trnsys[1:8760*2+1], atol=1e-3)  

if __name__ == "__main__":
    pytest.main()
